# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from datetime import datetime, time
import os
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# --- ⚙️ Page Config (Centered) ---
st.set_page_config(layout="centered", page_title="Production Dashboard", page_icon="🌌")

# --- 🎨 MINIMALIST LIGHT THEME CSS ---
st.markdown(
    """
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Prompt:wght@300;400;500;600;700;800&display=swap');
        
        :root {
            --bg-main: #f5f5f7;
            --bg-card: #ffffff;
            --border-color: #e0e0e4;
            --text-dark: #1f2937;
            --text-normal: #374151;
            --text-muted: #9ca3af;
            --accent-primary: #1e3a8a;
            --accent-light: #3b82f6;
            --header-light: #f3f4f6;
        }

        html, body, [class*="css"] {
            font-family: 'Prompt', sans-serif;
            color: var(--text-normal);
        }
        
        .stApp {
            background-color: var(--bg-main);
            background-attachment: fixed;
        }

        h1 {
            color: var(--accent-primary) !important;
            font-weight: 800 !important;
            font-size: 3.5rem !important;
            text-align: center;
            padding-bottom: 30px;
            letter-spacing: -1px;
            text-shadow: 1px 1px 5px rgba(0,0,0,0.05);
        }
        
        h3 {
            color: var(--text-dark) !important;
            font-size: 1.5rem !important;
            font-weight: 700 !important;
            margin-top: 15px;
            margin-bottom: 15px;
        }
        
        .stMarkdown p, .stcaption {
             color: var(--text-normal) !important;
             font-size: 1.0rem;
        }

        /* Tabs Styling */
        .stTabs [data-baseweb="tab-list"] {
            gap: 15px;
            background-color: var(--bg-card);
            padding: 10px;
            border-radius: 80px;
            border: 1px solid var(--border-color);
            margin-bottom: 40px;
            justify-content: center;
            box-shadow: 0 5px 15px rgba(0,0,0,0.05);
        }
        
        .stTabs [data-baseweb="tab"] {
            height: 65px;
            padding: 0 35px;
            border-radius: 50px;
            border: 1px solid transparent;
            background-color: transparent;
            transition: all 0.3s ease;
            color: var(--text-normal);
            font-family: 'Prompt', sans-serif;
            font-weight: 600; 
            font-size: 1.5rem; 
            letter-spacing: 0.5px;
            text-shadow: none;
        }
        
        .stTabs [data-baseweb="tab"]:hover {
            color: var(--accent-primary);
            background-color: var(--header-light);
            text-shadow: none;
        }

        .stTabs [aria-selected="true"] {
            background: var(--accent-primary);
            color: var(--bg-card) !important;
            font-weight: 800;
            box-shadow: 0 5px 15px rgba(30, 58, 138, 0.4);
            text-shadow: none;
        }

        /* Card Container */
        [data-testid="stVerticalBlock"] > [style*="flex-direction: column;"] > [data-testid="stVerticalBlock"] {
            background-color: var(--bg-card);
            border: 1px solid var(--border-color);
            border-radius: 20px;
            padding: 25px;
            box-shadow: 0 10px 20px -5px rgba(0, 0, 0, 0.05);
        }
        
        /* Buttons */
        .stButton > button {
            background: var(--accent-primary);
            color: var(--bg-card) !important;
            border: none;
            border-radius: 10px;
            font-weight: 700;
            font-size: 1rem;
            padding: 0.8rem 1.2rem;
            width: 100%;
            box-shadow: 0 4px 10px rgba(30, 58, 138, 0.4);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            transition: all 0.2s;
        }
        .stButton > button:hover {
             background-color: var(--accent-light);
             transform: translateY(-2px);
             box-shadow: 0 8px 20px rgba(60, 100, 200, 0.5);
        }

        /* DataFrame Styling */
        [data-testid="stDataFrame"] {
            background-color: var(--bg-card);
            border-radius: 12px;
            border: 1px solid var(--border-color);
            overflow: hidden;
        }
        [data-testid="stDataFrame"] th {
            background-color: var(--header-light) !important;
            color: var(--text-dark) !important;
            font-size: 0.95rem;
            font-weight: 600;
            border-bottom: 1px solid var(--border-color) !important;
        }
        [data-testid="stDataFrame"] td {
            color: var(--text-dark) !important;
            font-size: 0.95rem;
        }

        /* File Uploader */
        .stFileUploader {
            width: 100%;
            background-color: var(--bg-card) !important;
            border: 1px dashed var(--border-color) !important;
            border-radius: 12px;
            padding: 20px !important;
        }
        .stFileUploader:hover {
            border-color: var(--accent-light) !important;
            background-color: var(--header-light) !important;
        }
        .stFileUploader small { display: none; }
        [data-testid="stFileUploader"] button {
            padding: 0.2rem 0.8rem;
            font-size: 0.85rem;
        }
        
        /* Input fields */
        [data-baseweb="input"], [data-baseweb="select"], [data-testid="stDataEditor"] {
            background-color: var(--bg-card) !important;
            border: 1px solid var(--border-color) !important;
            color: var(--text-dark) !important;
            border-radius: 8px !important;
            font-weight: 500;
        }
        
        /* Expander */
        .streamlit-expanderHeader {
            background-color: var(--header-light) !important;
            color: var(--text-dark) !important;
            font-weight: 600;
            border-radius: 10px !important;
            border: 1px solid var(--border-color) !important;
        }
        .streamlit-expanderContent {
            background-color: var(--bg-card) !important;
            border: 1px solid var(--border-color);
            border-top: none;
            border-bottom-left-radius: 10px;
            border-bottom-right-radius: 10px;
            padding: 15px;
        }

        /* Animation */
        @keyframes slideUpFadeIn {
            0% { opacity: 0; transform: translateY(20px) scale(0.98); }
            100% { opacity: 1; transform: translateY(0) scale(1); }
        }
        [role="tabpanel"] > div {
            animation: slideUpFadeIn 0.5s cubic-bezier(0.2, 0.8, 0.2, 1) forwards;
        }
    </style>
    """,
    unsafe_allow_html=True
)

# --- Header ---
st.title("Percent Availability")

# --- Helper Functions ---
def validate_and_convert_hhmm(val):
    val_str = str(val).strip()
    if not re.match(r'^\d{1,2}:\d{1,2}$', val_str): return False, 0.0
    try:
        parts = val_str.split(":")
        h, m = int(parts[0]), int(parts[1])
        if m >= 60: return False, 0.0      
        return True, h + (m / 60)
    except: return False, 0.0

def get_shift_from_datetime(dt_val):
    try:
        dt = pd.to_datetime(dt_val, errors='coerce')
        if pd.isna(dt): return "Unknown"
        t = dt.time()
        if time(8, 0) <= t < time(20, 0):
            return "Day"
        else:
            return "Night"
    except:
        return "Unknown"

@st.cache_data(ttl=3600, show_spinner=False)
def load_capacity_file(uploaded_file):
    try:
        return pd.read_excel(uploaded_file, sheet_name="Capacity Counter")
    except:
        uploaded_file.seek(0) 
        return pd.read_excel(uploaded_file)

@st.cache_data(ttl=3600, show_spinner=False)
def load_job_file(uploaded_file):
    if uploaded_file.name.lower().endswith('.csv'):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip()
    return df

def find_column_by_keyword(columns, keywords):
    lower_cols = [str(c).lower().strip() for c in columns]
    for k in keywords:
        for i, col in enumerate(lower_cols):
            if k.lower() in col: return columns[i]
    return None

def format_seconds_to_hms(total_seconds):
    try:
        if pd.isna(total_seconds): return None
        ts = int(total_seconds)
        return f"{ts//3600:02d}:{(ts%3600)//60:02d}:{ts%60:02d}"
    except: return "00:00:00"

# ==========================================
# 🟢 MAIN TABS (Top Navigation)
# ==========================================
tab1, tab2, tab3, tab4 = st.tabs([
    "ZUND", 
    "SKIVE", 
    "BEAM PRESS", 
    "EXPORT"
])

# ==========================================
# TAB 1: Efficiency Calculator (ZUND)
# ==========================================
with tab1:
    if 'processing_stage' not in st.session_state: st.session_state['processing_stage'] = 'init'
    if 'data_loaded' not in st.session_state: st.session_state['data_loaded'] = False
    if 'df_import_state' not in st.session_state: st.session_state['df_import_state'] = None
    if 'df_capacity_data' not in st.session_state: st.session_state['df_capacity_data'] = None 
    if 'hours_setup_df' not in st.session_state: st.session_state['hours_setup_df'] = pd.DataFrame()
    if 'found_045_list' not in st.session_state: st.session_state['found_045_list'] = []
    
    # Input Card
    with st.container():
        st.markdown("### 📤 Upload Files")
        
        # ✅ Layout แบบ 2 คอลัมน์ (ซ้าย-ขวา)
        c1, c2 = st.columns(2)
        
        uploaded_capacity = c1.file_uploader(
            "1. Upload Capacity File", 
            type=['xlsx'], 
            key="capacity_ul", 
            label_visibility="visible"
        )
        
        uploaded_job = c2.file_uploader(
            "2. Zund File (Production)", 
            type=['xlsx', 'csv'], 
            key="job_ul", 
            label_visibility="visible"
        )
        
        st.write("") 
        
        if st.button("📥 READ DATA", type="primary", use_container_width=True):
            with st.spinner("🧑‍🏫 กำลังอ่านข้อมูล (Reading Data)..."):
                
                # 1. Reset ค่าระบบ
                st.session_state['hours_setup_df'] = pd.DataFrame()
                st.session_state['df_045_setup'] = pd.DataFrame()
                st.session_state['processing_stage'] = 'init'
                st.session_state['data_loaded'] = False
                st.session_state['found_045_list'] = [] 

                if not uploaded_job or not uploaded_capacity:
                    st.warning("⚠️ Please upload both Capacity and Job files.")
                    st.stop()
                
                try:
                    # 2. โหลดไฟล์
                    df_imp = load_job_file(uploaded_job)
                    st.session_state['df_import_state'] = df_imp
                    
                    df_cap = load_capacity_file(uploaded_capacity)
                    st.session_state['df_capacity_data'] = df_cap 
                    st.toast("✅ Files loaded.", icon="📂")

                    # =======================================================
                    # 🔴 LOGIC STRICT MODE: กรอง ZUND จากคอลัมน์ B และ E เท่านั้น
                    # =======================================================
                    final_list = []

                    # ตรวจสอบว่าไฟล์ Job มีอย่างน้อย 5 คอลัมน์ (Index 0 ถึง 4)
                    if len(df_imp.columns) >= 5:
                        
                        # --- STEP A: หา Part ที่มีปัญหา (K=0.45) จากไฟล์ Capacity ---
                        c_p_cap = find_column_by_keyword(df_cap.columns, ['Part No', 'Part']) or df_cap.columns[1]
                        bad_parts_all = []
                        if len(df_cap.columns) > 10:
                            col_k = df_cap.columns[10] # Column K (Index 10)
                            t_cap = df_cap[[c_p_cap, col_k]].copy()
                            t_cap.columns = ['Part', 'K']
                            t_cap['K'] = pd.to_numeric(t_cap['K'], errors='coerce').fillna(0)
                            # Part ทั้งหมดที่เป็น 0.45
                            bad_parts_all = t_cap[(t_cap['K'] - 0.45).abs() < 0.001]['Part'].astype(str).str.strip().tolist()

                        # --- STEP B: กรองไฟล์ Job (Force Column B & E) ---
                        # col_machine = Column Index 1 (B)
                        # col_part    = Column Index 4 (E)
                        col_machine = df_imp.iloc[:, 1].astype(str).str.strip().str.upper() 
                        col_part = df_imp.iloc[:, 4].astype(str).str.strip()             
                        
                        # สร้างเงื่อนไข: ต้องขึ้นต้นด้วย "ZUND" เท่านั้น (BEAM PRESS จะเป็น False)
                        is_zund_row = col_machine.str.startswith('ZUND')
                        
                        # ดึงชื่อ Part เฉพาะบรรทัดที่เป็น ZUND
                        parts_from_zund_only = col_part[is_zund_row].unique()
                        
                        # --- STEP C: Intersection (ตัวที่ซ้ำกัน) ---
                        final_list = [p for p in bad_parts_all if p in parts_from_zund_only]
                        
                        # แจ้งเตือนจำนวนแถวที่เจอ ZUND (Debug Info)
                        # st.toast(f"Found {is_zund_row.sum()} ZUND rows used for filtering.", icon="ℹ️")
                        
                    else:
                        st.error("❌ ไฟล์ Job Format ผิด: มีจำนวนคอลัมน์ไม่ถึง 5 คอลัมน์ (ต้องมีถึง Column E)")

                    # บันทึกผลลัพธ์
                    st.session_state['found_045_list'] = final_list
                    # =======================================================

                except Exception as e:
                    st.error(f"❌ Error: {e}")
                    st.stop()
                
                st.session_state['data_loaded'] = True
                
    # Process Card
    if st.session_state['data_loaded'] and st.session_state['df_import_state'] is not None:
        st.write("")
        with st.container():
            st.markdown("### ⚙️ Settings")
            df_import = st.session_state['df_import_state']
            df_capacity = st.session_state['df_capacity_data'] 
            
            # Map Columns Logic (For Calculation - Still uses flexible search for other calculations)
            if 'col_map' not in st.session_state or st.session_state['hours_setup_df'].empty:
                cols = df_import.columns
                c_p = find_column_by_keyword(cols, ['Part Number', 'Part', 'Model'])
                c_m = find_column_by_keyword(cols, ['Machine', 'Resource'])
                c_q = find_column_by_keyword(cols, ['Machine Qty', 'Qty'])
                c_t = find_column_by_keyword(cols, ['Date OUT', 'Time OUT', 'Date', 'Time'])
                
                if not all([c_p, c_m, c_q, c_t]): 
                    st.error("❌ Column mismatch in Job File (For Calculation).")
                else:
                    st.session_state['col_map'] = {'part': c_p, 'machine': c_m, 'qty': c_q, 'time': c_t}
                    st.session_state['machines'] = sorted(df_import[c_m].astype(str).unique())
                    
                    # Prepare Capacity Dictionary for Calculation
                    c_cp = find_column_by_keyword(df_capacity.columns, ['Part No', 'Part']) or df_capacity.columns[1]
                    if len(df_capacity.columns) > 11:
                         col_k = df_capacity.columns[10]
                         col_l = df_capacity.columns[11]
                         t_cap = df_capacity[[c_cp, col_k, col_l]].copy()
                         t_cap.columns = ['Part', 'K', 'L']
                         t_cap['K'] = pd.to_numeric(t_cap['K'], errors='coerce').fillna(0)
                         t_cap['L'] = pd.to_numeric(t_cap['L'], errors='coerce').fillna(1)
                         t_cap['Part'] = t_cap['Part'].astype(str).str.strip()
                         t_cap = t_cap.drop_duplicates(subset=['Part'], keep='first')
                         st.session_state['capacity_dict'] = t_cap.set_index('Part')[['K', 'L']].to_dict('index')
                    else:
                         st.error("❌ Capacity file structure error: Missing K/L columns.")
                         st.stop()

            # Set up DataFrame for the 0.45 Alert Table (Using the filtered list from strict logic)
            if st.session_state.get('found_045_list'):
                 st.session_state['df_045_setup'] = pd.DataFrame([{"Part": p, "New K": 0.45} for p in st.session_state['found_045_list']])
            else:
                 st.session_state['df_045_setup'] = pd.DataFrame()

            if st.session_state['hours_setup_df'].empty and 'machines' in st.session_state:
                st.session_state['hours_setup_df'] = pd.DataFrame([
                    {"MC": m, "Day": "08:00", "Night": "08:00", "Stat": "ปกติ"} 
                    for m in st.session_state['machines']
                ])

            # 4. K-Value Expander (Show only ZUND parts that are 0.45)
            edited_k_df = st.session_state['df_045_setup']
            if not edited_k_df.empty:
                with st.expander("⚠️ Part No without Cycletime (ZUND Only)", expanded=False):
                    edited_k_df = st.data_editor(
                        edited_k_df, 
                        use_container_width=True, 
                        hide_index=True, 
                        height=min(550, (len(edited_k_df) + 1) * 35 + 10), 
                        key="k_edit"
                    )
            
            btn_calc = False
            if not st.session_state['hours_setup_df'].empty:
                # 5. Machine Hours Collapsible
                with st.expander("⏱️ Machine Hours & Status Settings", expanded=False):
                    
                    full_df = st.session_state['hours_setup_df']
                    zund_mask = full_df['MC'].astype(str).str.upper().str.startswith('ZUND')
                    zund_df = full_df[zund_mask].copy()
                    
                    st.caption(f"Showing {len(zund_df)} ZUND machines.")
                    
                    edited_hours_zund = st.data_editor(
                        zund_df,
                        use_container_width=True,
                        height=min(550, (len(zund_df) + 1) * 35 + 10),
                        column_config={
                            "MC": st.column_config.TextColumn("Machine", disabled=True, width="small"),
                            "Day": st.column_config.TextColumn("Day (HH:MM)", width="small"),
                            "Night": st.column_config.TextColumn("Night (HH:MM)", width="small"),
                            "Stat": st.column_config.SelectboxColumn("Status", options=["ปกติ", "งานเศษ", "Sample"], width="small")
                        },
                        hide_index=True, key="h_edit_zund"
                    )

            # 6. Calculate Button
            st.write("")
            btn_calc = st.button("⚡ CALCULATE", type="primary", use_container_width=True)

            if btn_calc:
                final_calc_df = st.session_state['hours_setup_df'].copy()
                for idx, row in edited_hours_zund.iterrows():
                    mask = final_calc_df['MC'] == row['MC']
                    if mask.any():
                        final_calc_df.loc[mask, 'Day'] = row['Day']
                        final_calc_df.loc[mask, 'Night'] = row['Night']
                        final_calc_df.loc[mask, 'Stat'] = row['Stat']
                
                errs, c_day, c_night = [], [], []
                for i, r in final_calc_df.iterrows():
                    vd, vd_v = validate_and_convert_hhmm(r['Day'])
                    vn, vn_v = validate_and_convert_hhmm(r['Night'])
                    if not vd: errs.append(f"{r['MC']} Day Error")
                    if not vn: errs.append(f"{r['MC']} Night Error")
                    c_day.append(vd_v); c_night.append(vn_v)

                if errs: 
                    for e in errs: st.toast(f"❌ {e}")
                else:
                    st.session_state['hours_setup_df'] = final_calc_df 
                    calc_h = final_calc_df.copy()
                    calc_h['Day_F'] = c_day
                    calc_h['Night_F'] = c_night
                    
                    cur_map = st.session_state['capacity_dict'].copy()
                    if not edited_k_df.empty:
                        for _, r in edited_k_df.iterrows():
                            try:
                                new_k = float(r['New K'])
                                if str(r['Part']) in cur_map: cur_map[str(r['Part'])]['K'] = new_k
                            except: pass

                    cols = st.session_state['col_map']
                    df_c = df_import.copy()
                    df_c['P_Str'] = df_c[cols['part']].astype(str).str.strip()
                    try:
                        if len(df_import.columns) > 6: df_c['Raw_Output'] = pd.to_numeric(df_import.iloc[:, 6], errors='coerce').fillna(0)
                        else: df_c['Raw_Output'] = 0
                    except: df_c['Raw_Output'] = 0

                    df_c['K'] = df_c['P_Str'].map(lambda x: cur_map.get(x, {}).get('K', 0)).fillna(0)
                    df_c['L'] = df_c['P_Str'].map(lambda x: cur_map.get(x, {}).get('L', 1)).fillna(1).replace(0, 1)
                    df_c[cols['qty']] = pd.to_numeric(df_c[cols['qty']], errors='coerce').fillna(0)
                    df_c['Sec'] = (df_c[cols['qty']] / df_c['L']) * df_c['K'] * 60
                    df_c['Shift'] = df_c[cols['time']].apply(get_shift_from_datetime)
                    
                    piv = df_c.pivot_table(index=cols['machine'], columns="Shift", values=["Sec", "Raw_Output"], aggfunc="sum", fill_value=0).reset_index()
                    piv.columns = [f"{c[0]}_{c[1]}" if c[1] else c[0] for c in piv.columns.values]
                    piv.rename(columns={cols['machine']: "MC"}, inplace=True)
                    for col_name in ["Sec_Day", "Sec_Night", "Raw_Output_Day", "Raw_Output_Night"]:
                        if col_name not in piv.columns: piv[col_name] = 0.0
                    
                    merged = pd.merge(calc_h, piv, on="MC", how="left").fillna(0)
                    def get_disp(r):
                        s = r.get('Stat', 'ปกติ')
                        if s == 'งานเศษ': return f"{r['MC']} (⚠️)"
                        if s == 'Sample': return f"{r['MC']} (🧪)"
                        return r['MC']
                    merged['MC_Disp'] = merged.apply(get_disp, axis=1)
                    merged["Util_D"] = ((merged["Sec_Day"]/3600) / merged["Day_F"].replace(0,1)) * 100
                    merged["Util_N"] = ((merged["Sec_Night"]/3600) / merged["Night_F"].replace(0,1)) * 100
                    
                    norm = merged[merged['Stat'] == 'ปกติ']
                    avg_d = norm["Util_D"].mean() if not norm.empty else 0
                    avg_n = norm["Util_N"].mean() if not norm.empty else 0

                    final = pd.DataFrame({
                        "MC": merged["MC_Disp"],
                        "Day Time": merged["Sec_Day"].apply(format_seconds_to_hms),
                        "Day Output": merged["Raw_Output_Day"],
                        "Day%": merged["Util_D"],
                        "Night Time": merged["Sec_Night"].apply(format_seconds_to_hms),
                        "Night Output": merged["Raw_Output_Night"],
                        "Night%": merged["Util_N"]
                    })
                    
                    final = final[final['MC'].astype(str).str.lower() != 'nan']
                    final = final[final['MC'].notna()]
                    final = final[final['MC'].astype(str).str.strip() != '']

                    final = pd.concat([final, pd.DataFrame([{
                        "MC": "⭐ AVG", "Day Time": "-", "Day Output": "-", "Day%": avg_d, 
                        "Night Time": "-", "Night Output": "-", "Night%": avg_n
                    }])], ignore_index=True)
                    
                    target_bps = [f"BEAM PRESS {i}" for i in range(1, 6)] 
                    def clean_specific_machines(r):
                        mc_clean = str(r['MC']).split('(')[0].strip()
                        if mc_clean in target_bps:
                            r['Day Time'] = None; r['Day%'] = None
                            r['Night Time'] = None; r['Night%'] = None
                        return r
                    final = final.apply(clean_specific_machines, axis=1)

                    st.session_state['final_result'] = final
                    st.session_state['processing_stage'] = 'done'
                    st.rerun()

    # Result Card (Collapsible)
    if st.session_state.get('processing_stage') == 'done' and 'final_result' in st.session_state:
        st.write("")
        with st.expander("📊 Calculation Results", expanded=True):
            
            df_show = st.session_state['final_result'].copy()
            
            def clean_time(val):
                s = str(val).strip()
                if s == "00:00:00" or s.lower() == "nan" or s == "None" or not s: return "-"
                return s

            def clean_num(val, is_pct=False):
                try:
                    if pd.isna(val) or str(val).lower() == "nan" or val == "None": return "-"
                    f = float(val)
                    if f == 0: return "-"
                    if is_pct: return f"{f:.1f}%"
                    return f"{f:,.1f}"
                except: return "-"

            df_show["Day Time"] = df_show["Day Time"].apply(clean_time)
            df_show["Night Time"] = df_show["Night Time"].apply(clean_time)
            
            df_show["Day Output"] = df_show["Day Output"].apply(lambda x: clean_num(x))
            df_show["Night Output"] = df_show["Night Output"].apply(lambda x: clean_num(x))
            df_show["Day%"] = df_show["Day%"].apply(lambda x: clean_num(x, True))
            df_show["Night%"] = df_show["Night%"].apply(lambda x: clean_num(x, True))

            st.dataframe(
                df_show,
                use_container_width=True,
                height=(len(df_show) + 1) * 35 + 10,
                hide_index=True,
                column_config={
                    "MC": st.column_config.TextColumn("Machine", width="medium"),
                    "Day Time": st.column_config.TextColumn("Day Time", width="small"),
                    "Day Output": st.column_config.TextColumn("Day Qty", width="small"),
                    "Day%": st.column_config.TextColumn("Day Eff %", width="small"),
                    "Night Time": st.column_config.TextColumn("Night Time", width="small"),
                    "Night Output": st.column_config.TextColumn("Night Qty", width="small"),
                    "Night%": st.column_config.TextColumn("Night Eff %", width="small")
                }
            )

# ==========================================
# TAB 2: Skive Merge Tool
# ==========================================
with tab2:
    if 'skive_result_df' not in st.session_state: st.session_state['skive_result_df'] = None
    if 'skive_raw_df' not in st.session_state: st.session_state['skive_raw_df'] = None

    with st.container():
        st.markdown("### 📤 Upload File")
        
        c1, c2 = st.columns(2)
        sk_f1 = c1.file_uploader("1. Production File", type=['xlsx', 'csv'], key="sk1")
        sk_f2 = c2.file_uploader("2. Stock File", type=['xlsx', 'csv'], key="sk2")
        
        st.write("")
        # ✅ แก้ไข: ใช้เงื่อนไข OR (มีไฟล์เดียวก็ทำงานได้)
        if sk_f1 or sk_f2:
            if st.button("Merge Files", key="btn_skive", type="primary", use_container_width=True):
                try:
                    # ✅ เช็คทีละไฟล์ ถ้ามีก็โหลด ถ้าไม่มีก็สร้าง DataFrame ว่าง
                    df1 = load_job_file(sk_f1) if sk_f1 else pd.DataFrame()
                    df2 = load_job_file(sk_f2) if sk_f2 else pd.DataFrame()
                    
                    combined = pd.concat([df1, df2], ignore_index=True)
                    
                    if combined.empty:
                         st.warning("⚠️ No data found in uploaded files.")
                    else:
                        cols = combined.columns
                        
                        c_mc = find_column_by_keyword(cols, ['Machine', 'Resource', 'เครื่อง'])
                        if not c_mc and len(cols) > 1: c_mc = cols[1]
                        c_qty = find_column_by_keyword(cols, ['Machine Qty', 'Qty', 'จำนวน'])
                        if not c_qty and len(cols) > 5: c_qty = cols[5]
                        c_time = find_column_by_keyword(cols, ['Time', 'เวลา', 'Date', 'วันที่'])
                        if not c_time: c_time = cols[0] 
                        
                        if c_mc and c_qty:
                            ref_machines = pd.DataFrame({
                                'MC_Int': range(1, 16),
                                'MC_Show': [f"SKIVING {i:03d}" for i in range(1, 16)]
                            })
                            combined['MC_Int'] = combined[c_mc].astype(str).apply(
                                lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else -1
                            )
                            combined[c_qty] = pd.to_numeric(combined[c_qty], errors='coerce').fillna(0)
                            combined['Shift'] = combined[c_time].apply(get_shift_from_datetime)
                            
                            pivoted = combined.pivot_table(index='MC_Int', columns='Shift', values=c_qty, aggfunc='sum', fill_value=0).reset_index()
                            for s in ['Day', 'Night']:
                                if s not in pivoted.columns: pivoted[s] = 0
                            
                            final_skive = pd.merge(ref_machines, pivoted, on='MC_Int', how='left')
                            final_skive['Day'] = final_skive['Day'].fillna(0)
                            final_skive['Night'] = final_skive['Night'].fillna(0)
                            final_skive['Total'] = final_skive['Day'] + final_skive['Night']
                            
                            st.session_state['skive_raw_df'] = final_skive.copy()
                            def fmt(x): return f"{int(x):,}" if x > 0 else "-"
                            final_skive['Day_Show'] = final_skive['Day'].apply(fmt)
                            final_skive['Night_Show'] = final_skive['Night'].apply(fmt)
                            final_skive['Total_Show'] = final_skive['Total'].apply(fmt)
                            
                            st.session_state['skive_result_df'] = final_skive[['MC_Show', 'Day_Show', 'Night_Show', 'Total_Show']]
                            st.toast("✅ Merge Successful!", icon="🎉")
                        else:
                            st.error("❌ Column Missing")
                except Exception as e:
                    st.error(f"Error: {e}")

    if st.session_state['skive_result_df'] is not None:
        st.write("")
        with st.container():
            st.markdown("##### 📊 Skiving Summary")
            st.dataframe(
                st.session_state['skive_result_df'], 
                hide_index=True, use_container_width=True,
                height=(len(st.session_state['skive_result_df']) + 1) * 35 + 10
            )

# ==========================================
# TAB 3: Beam Press Summary
# ==========================================
with tab3:
    if 'bp_result_df' not in st.session_state: st.session_state['bp_result_df'] = None
    if 'bp_raw_df' not in st.session_state: st.session_state['bp_raw_df'] = None 

    with st.container():
        st.markdown("### 📤 Upload File")
        c1, c2 = st.columns(2)
        bp_f1 = c1.file_uploader("1. Production File", type=['xlsx', 'csv'], key="bp_prod")
        bp_f2 = c2.file_uploader("2. Stock File", type=['xlsx', 'csv'], key="bp_stock")

        st.write("")
        # ✅ แก้ไข: ใช้เงื่อนไข OR เช่นกัน
        if bp_f1 or bp_f2:
            if st.button("Merge Files", key="btn_bp", type="primary", use_container_width=True):
                try:
                    # ✅ เช็คทีละไฟล์
                    df1 = load_job_file(bp_f1) if bp_f1 else pd.DataFrame()
                    df2 = load_job_file(bp_f2) if bp_f2 else pd.DataFrame()
                    
                    combined = pd.concat([df1, df2], ignore_index=True)
                    
                    if combined.empty:
                        st.warning("⚠️ No data found in uploaded files.")
                    else:
                        cols = combined.columns

                        c_mc = find_column_by_keyword(cols, ['Machine', 'Resource', 'เครื่อง'])
                        if not c_mc: c_mc = cols[1] if len(cols) > 1 else cols[0] 
                        c_qty = find_column_by_keyword(cols, ['Machine Qty', 'Qty', 'จำนวน'])
                        if not c_qty: c_qty = cols[6] if len(cols) > 6 else cols[-1] 
                        c_time = find_column_by_keyword(cols, ['Time', 'เวลา', 'Date', 'วันที่'])
                        if not c_time: c_time = cols[0] 

                        combined[c_qty] = pd.to_numeric(combined[c_qty], errors='coerce').fillna(0)
                        def extract_mc_num(val):
                            s = str(val)
                            nums = re.findall(r'\d+', s)
                            return int(nums[-1]) if nums else -1
                            
                        combined['MC_Num'] = combined[c_mc].apply(extract_mc_num)
                        combined['Shift'] = combined[c_time].apply(get_shift_from_datetime)

                        target_machines = [1, 2, 3, 4, 5, 9, 10]
                        ref_df = pd.DataFrame({'MC_Num': target_machines})
                        ref_df['MC_Name'] = ref_df['MC_Num'].apply(lambda x: f"BEAM PRESS {x}")

                        pivoted = combined[combined['MC_Num'].isin(target_machines)].pivot_table(
                            index='MC_Num', columns='Shift', values=c_qty, 
                            aggfunc='sum', fill_value=0
                        ).reset_index()
                        
                        for s in ['Day', 'Night']:
                            if s not in pivoted.columns: pivoted[s] = 0
                        
                        final_bp = pd.merge(ref_df, pivoted, on='MC_Num', how='left')
                        final_bp['Day'] = final_bp['Day'].fillna(0)
                        final_bp['Night'] = final_bp['Night'].fillna(0)
                        final_bp['Total'] = final_bp['Day'] + final_bp['Night']

                        st.session_state['bp_raw_df'] = final_bp.copy()
                        def fmt(x): return f"{int(x):,}" if x > 0 else "-"
                        final_bp['Day_Show'] = final_bp['Day'].apply(fmt)
                        final_bp['Night_Show'] = final_bp['Night'].apply(fmt)
                        final_bp['Total_Show'] = final_bp['Total'].apply(fmt)

                        st.session_state['bp_result_df'] = final_bp[['MC_Name', 'Day_Show', 'Night_Show', 'Total_Show']]
                        st.toast("✅ Calculation Complete!", icon="🚜")

                except Exception as e:
                    st.error(f"❌ Error: {e}")

    if st.session_state['bp_result_df'] is not None:
        st.write("")
        with st.container():
            st.markdown("##### 📊 Beam Press Summary")
            st.dataframe(
                st.session_state['bp_result_df'],
                hide_index=True, use_container_width=True,
                height=(len(st.session_state['bp_result_df']) + 1) * 35 + 10
            )

# ==========================================
# TAB 4: Export to Report Template
# ==========================================
with tab4:
    
    EMBEDDED_TEMPLATE_FILENAME = "report_template.xlsx"
    
    st.write("---")
    st.markdown(f"### 🚀 Generate Report")

    # Check if data exists
    has_any_data = st.session_state.get('final_result') is not None or \
                   st.session_state.get('skive_raw_df') is not None or \
                   st.session_state.get('bp_raw_df') is not None
                   
    if st.button(
        "🚀 Fill Data & Generate Excel", 
        type="primary", 
        use_container_width=True,
        disabled=not has_any_data
    ):
        if not has_any_data:
            st.warning("⚠️ No calculation data found. Please run ZUND, SKIVE, or BEAM PRESS calculation first.")
        else:
            try:
                if not os.path.exists(EMBEDDED_TEMPLATE_FILENAME):
                    st.error(f"❌ Error: ไม่พบไฟล์ Template ชื่อ '{EMBEDDED_TEMPLATE_FILENAME}' ในโฟลเดอร์ Project.")
                    st.stop()
                    
                wb = load_workbook(EMBEDDED_TEMPLATE_FILENAME)
                ws = wb.active 
                
                df_eff = st.session_state.get('final_result')
                df_skive = st.session_state.get('skive_raw_df')
                df_bp = st.session_state.get('bp_raw_df')

                has_eff = df_eff is not None
                has_skive = df_skive is not None
                has_beam = df_bp is not None

                def write_val(sheet, r, c, val):
                    cell = sheet.cell(row=r, column=c)
                    try:
                        v_num = float(val)
                        if v_num == 0:
                            cell.value = "-"
                            cell.alignment = Alignment(horizontal='center')
                        else:
                            cell.value = v_num
                            cell.number_format = '#,##0.0'
                    except (ValueError, TypeError):
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal='center')

                def find_row(sheet, col_idx, search_text):
                    if not search_text: return None
                    search_text = str(search_text).lower().replace(" ", "").strip()
                    for row in range(1, 150):
                        cell_val = str(sheet.cell(row=row, column=col_idx).value).lower().replace(" ", "").strip()
                        if search_text in cell_val and cell_val != "none" and cell_val != "":
                            return row
                    return None
                
                # 1. Fill BEAM PRESS raw output
                if has_beam:
                    bp_summary_map = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 9: 8, 10: 9}
                    for _, row in df_bp.iterrows():
                        mc_num = row.get('MC_Num', -1)
                        if mc_num in bp_summary_map:
                            target_row = bp_summary_map[mc_num]
                            write_val(ws, target_row, 14, row['Day'])
                            write_val(ws, target_row, 15, row['Night'])

                # 2. Fill EFFICIENCY
                if has_eff:
                    eff_lookup = {}
                    for _, row in df_eff.iterrows():
                        clean_name = str(row['MC']).split('(')[0].strip()
                        eff_lookup[clean_name] = row

                    bp_eff_targets = {
                        "BEAM PRESS 1": 17, "BEAM PRESS 2": 18, "BEAM PRESS 3": 19, 
                        "BEAM PRESS 4": 20, "BEAM PRESS 5": 21, "BEAM PRESS 9": 22, "BEAM PRESS 10": 23
                    }

                    for mc_name, r_idx in bp_eff_targets.items():
                        if mc_name in eff_lookup:
                            row_data = eff_lookup[mc_name]
                            write_val(ws, r_idx, 14, row_data['Day Output'])
                            write_val(ws, r_idx, 15, row_data['Night Output'])
                        else:
                            write_val(ws, r_idx, 14, 0)
                            write_val(ws, r_idx, 15, 0)

                    # ZUND Time/Output
                    for _, row in df_eff.iterrows():
                        mc_str = str(row['MC'])
                        mc_name = mc_str.split("(")[0].strip() 
                        if "AVG" in mc_str or mc_name in bp_eff_targets: continue 

                        r_day = find_row(ws, 2, mc_name)
                        if r_day:
                            ws.cell(row=r_day, column=3).value = clean_time(row['Day Time'])
                            write_val(ws, r_day, 4, row['Day Output'])
                        
                        r_night = find_row(ws, 8, mc_name)
                        if r_night:
                            ws.cell(row=r_night, column=9).value = clean_time(row['Night Time'])
                            write_val(ws, r_night, 10, row['Night Output'])

                    # Write Average Efficiency
                    avg_row = df_eff[df_eff['MC'].astype(str).str.contains("AVG")]
                    if not avg_row.empty:
                        day_eff = avg_row['Day%'].values[0]
                        night_eff = avg_row['Night%'].values[0]
                        ws['D23'] = f"{day_eff:.2f}%"
                        ws['J23'] = f"{night_eff:.2f}%"

                # 3. Fill SKIVE output data
                if has_skive:
                    for _, row in df_skive.iterrows():
                        mc_name = str(row['MC_Show'])
                        try:
                            num = int(re.search(r'\d+', mc_name).group())
                            search_key = f"SKIVING {num}"
                        except:
                            search_key = mc_name
                        
                        r_sk = find_row(ws, 18, search_key)
                        if r_sk:
                            write_val(ws, r_sk, 19, row['Day'])
                            write_val(ws, r_sk, 20, row['Night'])

                # Save
                out_buffer = BytesIO()
                wb.save(out_buffer)
                
                st.toast("✅ Excel file generated successfully!", icon="🚀")
                st.download_button(
                    "📥 Download Result File", 
                    out_buffer.getvalue(), 
                    f"Filled_{EMBEDDED_TEMPLATE_FILENAME}", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheet",
                    type="primary"
                )
                
            except Exception as e:
                st.error(f"❌ Error during file generation: {e}")
